# utils.py
import openpyxl
from openpyxl.styles import Alignment
from oddsApi.settings import NFL_UNDERDOG_PROPS_FILE_PATH, NFL_PRIZEPICKS_PROPS_FILE_PATH
# markets => 
# [ player_pass_attempts, player_pass_completions, player_pass_interceptions, player_pass_longest_completion,
# player_pass_yds, player_pass_tds, player_reception_longest, player_receptions, player_reception_yds,
# player_rush_attempts, player_rush_longest, player_rush_reception_tds, player_rush_reception_yds, player_rush_yds,
# player_sacks, player_solo_tackles, player_tackles_assists, player_kicking_points, player_field_goals,
# player_defensive_interceptions, player_assists, player_pass_yds_q1 ]
# NFL_PLAYER_MARKETS = "player_pass_attempts,player_pass_completions,player_pass_interceptions,player_pass_longest_completion,player_pass_yds,player_pass_yds_q1,player_pass_tds,player_rush_attempts,player_rush_longest,player_rush_yds,player_rush_reception_yds,player_rush_reception_tds,player_reception_yds,player_reception_longest,player_receptions,player_sacks,player_solo_tackles,player_tackles_assists"
NFL_PLAYER_MARKETS = "player_rush_longest"
# alternate markets => 
# [ player_points_alternate, player_rebounds_alternate, player_assists_alternate, player_blocks_alternate	
# player_steals_alternate, player_turnovers_alternate, player_threes_alternate, player_points_assists_alternate,
# player_points_rebounds_alternate, player_rebounds_assists_alternate, player_points_rebounds_assists_alternate ]
NFL_ALTERNATE_PLAYER_MARKETS = ""
NFL_PLAYER_ODDS_REGIONS = "us,eu"
NFL_PLAYER_DFS_REGIONS = "us_dfs"
NFL_PLAYER_ODDS_FORMAT = "decimal"
NFL_BOOKMAKERS = ["pinnacle", "draftkings", "fanduel", "betmgm", "williamhill_us"]  # Filter list
NFL_BOOKMAKER_WEIGHTS = {
        "Pinnacle": 0.50,
        "Caesars": 0.20,
        "BetMGM": 0.20,
        "DraftKings": 0.05,
        "FanDuel": 0.05
    }


def filter_better_odds_lean(player_props, dfs_site):
    """
    Filters out the worse odds for each player by comparing 'over' and 'under' odds for each bookmaker.
    Keeps only the better (more negative) odds per bookmaker, based on the overall best lean.
    """
    filtered_props = []

    for prop in player_props:
        player_name = prop["player_name"]
        market = prop["market"]
        prop_line = prop["prop_line"]
        bookmaker_odds = prop["bookmaker_odds"]

        market = format_market(market.lower())
        
        bookmaker_groups = {}  # Store "over" and "under" odds together by bookmaker

        # Group odds by bookmaker
        for odds_entry in bookmaker_odds:
            bookmaker = odds_entry["bookmaker"]
            lean = odds_entry["lean"].lower()  # "over" or "under"
            odds = odds_entry["odds"]

            if bookmaker not in bookmaker_groups:
                bookmaker_groups[bookmaker] = {"over": None, "under": None}

            bookmaker_groups[bookmaker][lean] = odds

        # Calculate fair odds for each bookmaker
        fair_odds_per_bookmaker = []
        for bookmaker, odds_pair in bookmaker_groups.items():
            over_odds = odds_pair["over"]
            under_odds = odds_pair["under"]

            if over_odds and under_odds:  # Ensure both exist
                fair_over, fair_under = calculate_fair_odds(over_odds, under_odds)

                fair_odds_per_bookmaker.append({
                    "bookmaker": bookmaker,
                    "fair_over": fair_over,
                    "fair_under": fair_under
                })

        # Calculate weighted fair odds (only consider bookmakers that exist in the odds data)
        weighted_fair_over = 0
        weighted_fair_under = 0
        total_weight = 0

        for fair_odds in fair_odds_per_bookmaker:
            bookmaker = fair_odds["bookmaker"]

            # Check if the bookmaker exists in our weights list
            if bookmaker in NFL_BOOKMAKER_WEIGHTS:
                weight = NFL_BOOKMAKER_WEIGHTS[bookmaker]
                weighted_fair_over += fair_odds["fair_over"] * weight
                weighted_fair_under += fair_odds["fair_under"] * weight
                total_weight += weight
            
        # Normalize by the total weight (if total_weight > 0, to prevent division by zero)
        avg_fair_over_odds = weighted_fair_over / total_weight if total_weight > 0 else None
        avg_fair_under_odds = weighted_fair_under / total_weight if total_weight > 0 else None

        avg_fair_over_odds = round(avg_fair_over_odds, 2) if avg_fair_over_odds is not None else None
        avg_fair_under_odds = round(avg_fair_under_odds, 2) if avg_fair_under_odds is not None else None

        # Determine better lean based on fair odds averages
        better_lean = ""
        if avg_fair_over_odds and avg_fair_under_odds:
            if avg_fair_over_odds < avg_fair_under_odds:
                better_lean = "over"
            elif avg_fair_over_odds > avg_fair_under_odds:
                better_lean = "under"
            else:
                better_lean = "n/a"
        elif avg_fair_over_odds:  # If there's no under odds, pick over
            better_lean = "over"
        elif avg_fair_under_odds:  # If there's no over odds, pick under
            better_lean = "under"

        if better_lean == "n/a":
            continue

        # Filter bookmaker odds for the better lean
        selected_bookmaker_odds = []
        for odds_entry in bookmaker_odds:
            if odds_entry["lean"].lower() == better_lean:
                selected_bookmaker_odds.append({
                    "lean": odds_entry["lean"],
                    "market": market,
                    "point": prop_line,
                    "odds": decimal_to_american(odds_entry["odds"]),
                    "bookmaker": odds_entry["bookmaker"]
                })

        # Determine the average odds
        average_odds = decimal_to_american(avg_fair_over_odds) if better_lean == "over" else decimal_to_american(avg_fair_under_odds)

        # Check if at least two bookmakers exist, and one of them must be Pinnacle, BetMGM, or Caesars
        if len(bookmaker_groups) >= 2 and any(bookmaker in ["Pinnacle", "BetMGM", "Caesars"] for bookmaker in bookmaker_groups):
            filtered_props.append({
                "player_name": player_name,
                "market": market,
                "point": prop_line,
                "lean": better_lean,
                "average_fair_odds": average_odds,
                "fair_probability": implied_probability(average_odds),
                "bookmaker_odds": selected_bookmaker_odds
            })


        # Sort by implied probability in descending order
        filtered_props.sort(key=lambda x: x["fair_probability"], reverse=True)

        # Keep only the top 20 elements if there are more than 20
        filtered_props = filtered_props[:15]

        # Export to the respective site
        if dfs_site.lower() == "ud":
            export_to_excel(filtered_props, NFL_UNDERDOG_PROPS_FILE_PATH)
        elif dfs_site.lower() == "pp":
            export_to_excel(filtered_props, NFL_PRIZEPICKS_PROPS_FILE_PATH)

    return filtered_props


def decimal_to_american(decimal_odds):
    if decimal_odds >= 2.0:
        # Positive American odds
        american_odds = (decimal_odds - 1) * 100
    else:
        # Negative American odds
        american_odds = - (100 / (decimal_odds - 1))
    
    return round(american_odds)


def format_market(market):
    switcher = {
        "player_pass_attempts": "Pass Attempts",
        "player_pass_completions": "Pass Completions",
        "player_pass_interceptions": "Pass Ints",
        "player_pass_longest_completion": "Longest Completion",
        "player_pass_yds": "Pass Yards",
        "player_pass_tds": "Pass TDs",
        "player_reception_longest": "Longest Reception",
        "player_receptions": "Receptions",
        "player_reception_yds": "Receiving Yards",
        "player_rush_attempts": "Rush Attempts",
        "player_rush_longest": "Longest Rush",
        "player_rush_reception_tds": "Rush+Rec TDs",
        "player_rush_reception_yds": "Rush+Rec Yards",
        "player_rush_yds": "Rushing Yards",
        "player_sacks": "Sacks",
        "player_solo_tackles": "Solo Tackles",
        "player_tackles_assists": "Tackles+Assists",
        "player_kicking_points": "Kicking Points",
        "player_field_goals": "Field Goals",
        "player_defensive_interceptions": "Def Ints",
        "player_assists": "Assists",
        "player_pass_yds_q1": "Pass Yards 1Q"
    }
    return switcher.get(market, "unknown")  # Default to "unknown" if market doesn't match any case

def implied_probability(average_odds):
    """
    Calculate the implied probability from the given average odds.
    Works for both positive and negative American odds.
    
    Args:
        average_odds (float): The average odds in American format.
    
    Returns:
        float: The implied probability, as a percentage.
    """
    if average_odds > 0:  # Positive odds
        probability = 100 / (average_odds + 100)
    else:  # Negative odds
        probability = -average_odds / (-average_odds + 100)
    
    return round(probability * 100, 2)  # Return as percentage, rounded to 2 decimal places


def export_to_excel(data, filename):
    # Define the headers as you requested
    headers = ["Player Name", "Lean", "Prop Line", "Market", "Pinnacle", "Caesars", "BetMGM", "DraftKings", "FanDuel", "Fair Probability"]

    try:
        # Create a new workbook and active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Player Props"

        # Add the headers to the first row and center align them
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
            # Apply center alignment for headers
            sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

        # Loop through the data and add rows for each prop
        for row_num, prop in enumerate(data, 2):
            player_name = prop.get("player_name", "n/a")
            market = prop.get("market", "n/a")
            lean = prop.get("lean", "n/a")
            prop_line = prop.get("point", "n/a")
            fair_probability = prop.get("fair_probability", "n/a")
            bookmaker_odds = prop.get("bookmaker_odds", [])

            # combined_prop_line_market = f"{prop_line} {market}"

            # Initialize the odds dictionary with "n/a" for each bookmaker
            bookmaker_columns = {
                "Pinnacle": "",
                "Caesars": "",
                "BetMGM": "",
                "DraftKings": "",
                "FanDuel": "",
            }

            # Loop through each bookmaker and add the odds to the correct column
            for odds_entry in bookmaker_odds:
                bookmaker_name = odds_entry["bookmaker"]
                if bookmaker_name in bookmaker_columns:
                    bookmaker_columns[bookmaker_name] = odds_entry["odds"]

            # Flatten the odds and implied probabilities for the row
            row = [player_name, lean.capitalize(), prop_line, market] + list(bookmaker_columns.values()) + [fair_probability]

            # Add the row data to the sheet and apply center alignment
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = Alignment(horizontal="center")

        # Adjust column width based on the longest cell content
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            for row_num in range(1, len(data) + 1):  # +2 for headers and rows
                cell_value = str(sheet.cell(row=row_num, column=col_num).value)
                max_length = max(max_length, len(cell_value))
            adjusted_width = (max_length + 2)  # Add a little padding
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width

        # Save the workbook to the specified filename
        wb.save(filename)

        print(f"Data successfully exported")

    except Exception as e:
        print(f"Error exporting data: {e}")


def calculate_fair_odds(over_odds, under_odds):
    """
    Given decimal odds for 'over' and 'under', calculate the fair odds by removing the vig.
    """
    # Convert Decimal odds to Implied Probability
    over_prob = 1 / over_odds
    under_prob = 1 / under_odds

    # Total implied probability with vig
    total_prob = over_prob + under_prob

    # Normalize probabilities to remove the vig
    fair_over_prob = over_prob / total_prob
    fair_under_prob = under_prob / total_prob

    # Convert fair probabilities back to Decimal odds
    fair_over_odds = 1 / fair_over_prob if fair_over_prob != 0 else None
    fair_under_odds = 1 / fair_under_prob if fair_under_prob != 0 else None

    return fair_over_odds, fair_under_odds