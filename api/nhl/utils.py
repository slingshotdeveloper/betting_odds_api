# utils.py
import openpyxl
from openpyxl.styles import Alignment
from oddsApi.settings import NHL_UNDERDOG_PROPS_FILE_PATH, NHL_PRIZEPICKS_PROPS_FILE_PATH
# markets => 
# [ player_points, player_assists, player_blocked_shots, player_shots_on_goal, player_goals, player_total_saves ]
NHL_PLAYER_MARKETS = "player_points"
NHL_PLAYER_ODDS_REGIONS = "us"
NHL_PLAYER_DFS_REGIONS = "us_dfs"
NHL_PLAYER_ODDS_FORMAT = "decimal"

def filter_better_odds_selection(player_props, dfs_site):
    """
    Filters out the worse odds for each player by comparing 'over' and 'under' odds for each bookmaker.
    Keeps only the better (more negative) odds per bookmaker, based on the overall best selection.
    """
    filtered_props = []

    for prop in player_props:
        player_name = prop["player_name"]
        market = prop["market"]
        prop_line = prop["prop_line"]
        bookmaker_odds = prop["bookmaker_odds"]

        market = format_market(market.lower())
        
        over_odds_total = 0
        under_odds_total = 0
        over_count = 0
        under_count = 0
        better_selection = ""

        # Organize odds by bookmaker
        for odds_entry in bookmaker_odds:
            selection = odds_entry["selection"].lower()  # "over" or "under"
            odds = odds_entry["odds"]

            # Track sum and count for averaging
            if selection == "over":
                over_odds_total += odds
                over_count += 1
            elif selection == "under":
                under_odds_total += odds
                under_count += 1

        # Calculate average odds, avoid division by zero
        avg_over_odds = over_odds_total / over_count if over_count > 0 else None
        avg_under_odds = under_odds_total / under_count if under_count > 0 else None

        # Round average odds to two decimal places
        avg_over_odds = round(avg_over_odds, 2) if avg_over_odds is not None else None
        avg_under_odds = round(avg_under_odds, 2) if avg_under_odds is not None else None

        # Determine better selection based on averages
        if avg_over_odds and avg_under_odds:
            if avg_over_odds < avg_under_odds:
                better_selection = "over"
            elif avg_over_odds > avg_under_odds:
                better_selection = "under"
            else:
                better_selection = "n/a"
        elif avg_over_odds:  # If there's no under odds, pick over
            better_selection = "over"
        elif avg_under_odds:  # If there's no over odds, pick under
            better_selection = "under"

        if better_selection == "n/a":
            continue
        
        # Filter bookmaker odds for the better selection
        selected_bookmaker_odds = []
        for odds_entry in bookmaker_odds:
            if odds_entry["selection"].lower() == better_selection:
                selected_bookmaker_odds.append({
                    "selection": odds_entry["selection"],
                    "market": market,
                    "point": prop_line,
                    "odds": decimal_to_american(odds_entry["odds"]),
                    "bookmaker": odds_entry["bookmaker"]
                })

        average_odds = decimal_to_american(avg_over_odds) if better_selection == "over" else decimal_to_american(avg_under_odds)

        # Add the filtered prop to the result list if there are more than 2 bookmaker odds
        if len(selected_bookmaker_odds) > 2:
            filtered_props.append({
                "player_name": player_name,
                "market": market,
                "point": prop_line,
                "lean": better_selection,
                "average_odds": average_odds,
                "implied_probability": implied_probability(average_odds),
                "bookmaker_odds": selected_bookmaker_odds
            })

        filtered_props.sort(key=lambda x: x["implied_probability"], reverse=True)

        # Keep only the top 20 elements if there are more than 20
        filtered_props = filtered_props[:20]

        if dfs_site.lower() == "ud":
            export_to_excel(filtered_props, NHL_UNDERDOG_PROPS_FILE_PATH)
        elif dfs_site.lower() == "pp":
            export_to_excel(filtered_props, NHL_PRIZEPICKS_PROPS_FILE_PATH)

    return filtered_props


def decimal_to_american(decimal_odds):
    if decimal_odds >= 2.0:
        # Positive American odds
        american_odds = (decimal_odds - 1) * 100
    else:
        # Negative American odds
        american_odds = - (100 / (decimal_odds - 1))
    
    return round(american_odds)

def format_market(market):
    switcher = {
        "player_points": "Points",
        "player_assists": "Assists",
        "player_blocked_shots": "Blocked Shots",
        "player_shots_on_goal": "Shots on Goal",
        "player_goals": "Goals",
        "player_total_saves": "Saves"
    }
    return switcher.get(market, "unknown")  # Default to "unknown" if market doesn't match any case

def implied_probability(average_odds):
    """
    Calculate the implied probability from the given average odds.
    Works for both positive and negative American odds.
    
    Args:
        average_odds (float): The average odds in American format.
    
    Returns:
        float: The implied probability, as a percentage.
    """
    if average_odds > 0:  # Positive odds
        probability = 100 / (average_odds + 100)
    else:  # Negative odds
        probability = -average_odds / (-average_odds + 100)
    
    return round(probability * 100, 2)  # Return as percentage, rounded to 2 decimal places

def export_to_excel(data, filename):
    # Define the headers as you requested
    headers = ["Player Name", "Lean", "Prop Line", "Market", "DraftKings", "FanDuel", "BetRivers", "BetOnline.ag", "Bovada", "BetMGM", "Implied Probability"]

    try:
        # Create a new workbook and active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Player Props"

        # Add the headers to the first row and center align them
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
            # Apply center alignment for headers
            sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

        # Loop through the data and add rows for each prop
        for row_num, prop in enumerate(data, 2):
            player_name = prop.get("player_name", "n/a")
            market = prop.get("market", "n/a")
            lean = prop.get("lean", "n/a")
            prop_line = prop.get("point", "n/a")
            implied_probability = prop.get("implied_probability", "n/a")
            bookmaker_odds = prop.get("bookmaker_odds", [])

            # combined_prop_line_market = f"{prop_line} {market}"

            # Initialize the odds dictionary with "n/a" for each bookmaker
            bookmaker_columns = {
                "DraftKings": "n/a",
                "FanDuel": "n/a",
                "BetRivers": "n/a",
                "BetOnline.ag": "n/a",
                "Bovada": "n/a",
                "BetMGM": "n/a"
            }

            # Loop through each bookmaker and add the odds to the correct column
            for odds_entry in bookmaker_odds:
                bookmaker_name = odds_entry["bookmaker"]
                if bookmaker_name in bookmaker_columns:
                    bookmaker_columns[bookmaker_name] = odds_entry["odds"]

            # Flatten the odds and implied probabilities for the row
            row = [player_name, lean.capitalize(), prop_line, market] + list(bookmaker_columns.values()) + [implied_probability]

            # Add the row data to the sheet and apply center alignment
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = Alignment(horizontal="center")

        # Adjust column width based on the longest cell content
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            for row_num in range(1, len(data) + 1):  # +2 for headers and rows
                cell_value = str(sheet.cell(row=row_num, column=col_num).value)
                max_length = max(max_length, len(cell_value))
            adjusted_width = (max_length + 2)  # Add a little padding
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width

        # Save the workbook to the specified filename
        wb.save(filename)

        print(f"Data successfully exported")

    except Exception as e:
        print(f"Error exporting data: {e}")